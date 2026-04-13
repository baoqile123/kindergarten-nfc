#!/usr/bin/env python3
"""
幼儿园NFC签到系统 - Excel链接更新工具
将Excel表格中的freekit.dev链接更新为GitHub Pages链接
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from pathlib import Path
import json

def print_header():
    """打印标题"""
    print("=" * 70)
    print("📊 Excel链接更新工具")
    print("将freekit.dev链接更新为GitHub Pages链接")
    print("=" * 70)
    print()

def get_github_info():
    """获取GitHub信息"""
    print("📋 请提供GitHub信息以生成新链接：")
    
    username = input("GitHub用户名: ").strip()
    if not username:
        print("❌ 用户名不能为空")
        exit(1)
    
    repo_name = input("仓库名称（默认: kindergarten-nfc）: ").strip()
    if not repo_name:
        repo_name = "kindergarten-nfc"
    
    base_url = f"https://{username}.github.io/{repo_name}/nfc-checkin.html"
    
    print(f"\n✅ 新链接格式：")
    print(f"   基础链接：{base_url}")
    print(f"   手环模板：{base_url}#姓名")
    print(f"   区域链接：{base_url}?area=区域名")
    
    return base_url

def load_excel_files():
    """加载Excel文件列表"""
    print("\n📁 扫描Excel文件...")
    
    excel_files = []
    for file in Path(".").glob("*.xlsx"):
        excel_files.append(file.name)
    
    if not excel_files:
        print("❌ 未找到Excel文件")
        return []
    
    print(f"✅ 找到 {len(excel_files)} 个Excel文件：")
    for i, file in enumerate(excel_files, 1):
        print(f"   {i}. {file}")
    
    return excel_files

def update_excel_file(filepath, base_url):
    """更新单个Excel文件"""
    print(f"\n🔄 处理文件：{filepath}")
    
    try:
        # 读取Excel文件
        df = pd.read_excel(filepath)
        original_shape = df.shape
        print(f"   原始数据：{original_shape[0]}行 × {original_shape[1]}列")
        
        # 查找包含链接的列
        old_link = "https://freekit.dev/s/ejyVDEKU"
        link_columns = []
        
        for col in df.columns:
            if df[col].astype(str).str.contains(old_link, na=False).any():
                link_columns.append(col)
                print(f"   ✅ 发现链接列：{col}")
        
        if not link_columns:
            print("   ⚠️  未找到包含旧链接的列")
            return False
        
        # 更新链接
        updated_count = 0
        for col in link_columns:
            # 统计更新数量
            mask = df[col].astype(str).str.contains(old_link, na=False)
            count = mask.sum()
            
            if count > 0:
                # 替换链接
                df[col] = df[col].astype(str).str.replace(
                    old_link, 
                    base_url,
                    regex=False
                )
                updated_count += count
                print(f"   📝 更新 {col} 列：{count} 条记录")
        
        # 保存更新后的Excel文件（带格式）
        save_with_format(filepath, df, base_url, updated_count)
        
        print(f"   ✅ 更新完成：{updated_count} 条链接已更新")
        return True
        
    except Exception as e:
        print(f"   ❌ 处理失败：{e}")
        return False

def save_with_format(original_path, df, base_url, updated_count):
    """保存带格式的Excel文件"""
    # 创建新文件名
    new_filename = original_path.stem + "_GitHub版.xlsx"
    new_path = original_path.parent / new_filename
    
    # 先用pandas保存数据
    df.to_excel(new_path, index=False)
    
    # 用openpyxl添加格式
    wb = openpyxl.load_workbook(new_path)
    ws = wb.active
    
    # 设置标题样式
    header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # 设置数据样式
    data_font = Font(name="微软雅黑", size=10)
    link_font = Font(name="微软雅黑", size=10, color="0563C1", underline="single")
    highlight_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    
    # 应用标题样式
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # 查找并高亮链接列
    for col in range(1, ws.max_column + 1):
        col_name = ws.cell(row=1, column=col).value
        if col_name and isinstance(col_name, str) and ("链接" in col_name or "URL" in col_name):
            print(f"   🎨 格式化链接列：{col_name}")
            
            # 高亮列标题
            header_cell = ws.cell(row=1, column=col)
            header_cell.fill = PatternFill(start_color="C5D9F1", end_color="C5D9F1", fill_type="solid")
            
            # 格式化链接内容
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value and base_url in str(cell.value):
                    cell.font = link_font
                    cell.fill = highlight_fill
    
    # 添加信息页
    add_info_sheet(wb, base_url, updated_count, original_path.name)
    
    # 调整列宽
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    # 保存文件
    wb.save(new_path)
    print(f"   💾 保存为：{new_filename}（带格式）")

def add_info_sheet(wb, base_url, updated_count, original_filename):
    """添加信息说明页"""
    if "信息说明" in wb.sheetnames:
        ws_info = wb["信息说明"]
    else:
        ws_info = wb.create_sheet(title="信息说明")
    
    # 设置信息内容
    info_data = [
        ["📋 Excel链接更新说明"],
        [],
        ["更新信息"],
        ["原始文件：", original_filename],
        ["更新日期：", "2026-04-13"],
        ["更新链接数：", updated_count],
        ["GitHub Pages链接：", base_url],
        [],
        ["🏫 6个运动区域"],
        ["区域名称", "链接格式", "示例"],
        ["灌篮小乐宝", f"{base_url}?area=灌篮小乐宝", f"{base_url}?area=灌篮小乐宝"],
        ["IT小乐宝", f"{base_url}?area=IT小乐宝", f"{base_url}?area=IT小乐宝"],
        ["智驾小乐宝", f"{base_url}?area=智驾小乐宝", f"{base_url}?area=智驾小乐宝"],
        ["智酷小乐宝", f"{base_url}?area=智酷小乐宝", f"{base_url}?area=智酷小乐宝"],
        ["绳趣小乐宝", f"{base_url}?area=绳趣小乐宝", f"{base_url}?area=绳趣小乐宝"],
        ["芯动小乐宝", f"{base_url}?area=芯动小乐宝", f"{base_url}?area=芯动小乐宝"],
        [],
        ["📱 手环链接格式"],
        ["模板：", f"{base_url}#姓名"],
        ["示例：", f"{base_url}#冯珺羿"],
        [],
        ["🔑 键盘快捷键"],
        ["A键", "切换区域"],
        ["E键", "导出CSV记录"],
        [],
        ["💡 使用说明"],
        ["1. 将此Excel文件发送给相关老师"],
        ["2. 使用NFC Tools APP写入手环"],
        ["3. 平板浏览器打开对应区域链接"],
        ["4. 手环触碰平板即可签到"],
        [],
        ["📞 技术支持"],
        ["如有问题，请参考《迁移操作流程_完整版.md》"],
        ["或联系幼儿园技术负责人"]
    ]
    
    # 写入数据
    for row_idx, row_data in enumerate(info_data, 1):
        for col_idx, cell_data in enumerate(row_data, 1):
            cell = ws_info.cell(row=row_idx, column=col_idx, value=cell_data)
            
            # 设置标题样式
            if row_idx == 1:
                cell.font = Font(name="微软雅黑", size=14, bold=True, color="4F81BD")
            elif row_data and isinstance(row_data[0], str) and ("🏫" in row_data[0] or "📱" in row_data[0] or "🔑" in row_data[0] or "💡" in row_data[0] or "📞" in row_data[0]):
                cell.font = Font(name="微软雅黑", size=12, bold=True, color="2E75B6")
            elif row_idx == 3:  # "更新信息"标题
                cell.font = Font(name="微软雅黑", size=12, bold=True, color="2E75B6")
            elif col_idx == 1 and row_idx >= 4 and row_idx <= 7:  # 更新信息标签
                cell.font = Font(name="微软雅黑", size=11, bold=True)
                cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            elif col_idx == 2 and row_idx >= 4 and row_idx <= 7:  # 更新信息值
                cell.font = Font(name="微软雅黑", size=11)
            
            # 设置链接样式
            if cell_data and isinstance(cell_data, str) and base_url in cell_data:
                cell.font = Font(name="微软雅黑", size=10, color="0563C1", underline="single")
                cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    
    # 调整列宽
    ws_info.column_dimensions['A'].width = 25
    ws_info.column_dimensions['B'].width = 60
    ws_info.column_dimensions['C'].width = 60
    
    # 冻结首行
    ws_info.freeze_panes = "A2"

def create_summary_report(excel_files, base_url, results):
    """创建更新总结报告"""
    print("\n📊 生成更新总结报告...")
    
    summary = {
        "更新日期": "2026-04-13",
        "GitHub链接": base_url,
        "处理文件数": len(excel_files),
        "成功更新": sum(1 for r in results if r[1]),
        "失败文件": [file for file, success in results if not success],
        "文件详情": []
    }
    
    for i, (filepath, success) in enumerate(results, 1):
        summary["文件详情"].append({
            "序号": i,
            "文件名": filepath,
            "状态": "✅ 成功" if success else "❌ 失败",
            "新文件名": f"{Path(filepath).stem}_GitHub版.xlsx" if success else "未生成"
        })
    
    # 保存JSON报告
    report_file = "Excel更新报告.json"
    with open(report_file, 'w', encoding='utf-8') as f:
        json.dump(summary, f, ensure_ascii=False, indent=2)
    
    # 创建Markdown报告
    md_report = f"""# 📊 Excel链接更新报告

## 更新概要
- **更新日期**：2026-04-13
- **GitHub Pages链接**：{base_url}
- **处理文件数**：{len(excel_files)} 个
- **成功更新**：{sum(1 for r in results if r[1])} 个
- **失败文件**：{len([file for file, success in results if not success])} 个

## 文件处理详情

| 序号 | 文件名 | 状态 | 新文件名 |
|:---|:---|:---|:---|
"""
    
    for i, (filepath, success) in enumerate(results, 1):
        status = "✅ 成功" if success else "❌ 失败"
        new_filename = f"`{Path(filepath).stem}_GitHub版.xlsx`" if success else "未生成"
        md_report += f"| {i} | `{filepath}` | {status} | {new_filename} |\n"
    
    md_report += f"""
## 新链接格式

### 手环链接模板
```
{base_url}#姓名
```

### 示例手环链接
```
{base_url}#冯珺羿
{base_url}#李小明
{base_url}#张小红
```

### 6个运动区域链接
1. **🏀 灌篮小乐宝**：`{base_url}?area=灌篮小乐宝`
2. **💻 IT小乐宝**：`{base_url}?area=IT小乐宝`
3. **🚗 智驾小乐宝**：`{base_url}?area=智驾小乐宝`
4. **🎯 智酷小乐宝**：`{base_url}?area=智酷小乐宝`
5. **🪢 绳趣小乐宝**：`{base_url}?area=绳趣小乐宝`
6. **🧗 芯动小乐宝**：`{base_url}?area=芯动小乐宝`

## 使用说明

### 教师操作
1. 打开新生成的Excel文件（带"_GitHub版"后缀）
2. 使用"信息说明"页查看详细说明
3. 按班级分发手环链接信息
4. 使用NFC Tools APP写入手环

### 平板设置
每个运动区域的平板应打开对应的区域链接：
- 灌篮小乐宝平板：`{base_url}?area=灌篮小乐宝`
- IT小乐宝平板：`{base_url}?area=IT小乐宝`
- ...（其他区域同理）

### 数据管理
- 签到数据存储在平板本地
- 按**E键**可导出CSV记录
- 定期备份签到数据

## 注意事项

### ✅ 已完成
- 所有Excel文件链接已更新
- 生成带格式的新文件
- 添加详细使用说明页

### ⚠️ 需注意
- 请使用新文件替换旧文件
- 确保手环写入新链接格式
- 平板浏览器更新收藏链接

### 📞 技术支持
如有问题，请参考：
- 《迁移操作流程_完整版.md》
- 《验证总结报告.md》
- 联系幼儿园技术负责人

---
**报告生成时间**：2026-04-13  
**更新负责人**：幼儿园技术团队</summary>"""
    
    md_file = "Excel更新报告.md"
    with open(md_file, 'w', encoding='utf-8') as f:
        f.write(md_report)
    
    print(f"✅ JSON报告：{report_file}")
    print(f"✅ Markdown报告：{md_file}")
    
    # 显示简要总结
    print("\n📋 更新结果摘要：")
    print(f"   总文件数：{len(excel_files)}")
    print(f"   成功更新：{sum(1 for r in results if r[1])}")
    print(f"   失败文件：{len([file for file, success in results if not success])}")
    
    if any(not success for _, success in results):
        print("\n⚠️  有文件更新失败，请检查：")
        for filepath, success in results:
            if not success:
                print(f"   ❌ {filepath}")

def main():
    """主函数"""
    print_header()
    
    # 获取GitHub信息
    base_url = get_github_info()
    
    # 加载Excel文件
    excel_files = load_excel_files()
    if not excel_files:
        return
    
    # 确认继续
    print("\n⚠️  确认更新所有Excel文件吗？")
    confirm = input("输入 'yes' 继续，其他任意键取消：").strip().lower()
    if confirm != 'yes':
        print("❌ 操作已取消")
        return
    
    # 处理每个Excel文件
    results = []
    for filepath in excel_files:
        success = update_excel_file(filepath, base_url)
        results.append((filepath, success))
    
    # 生成总结报告
    create_summary_report(excel_files, base_url, results)
    
    print("\n" + "=" * 70)
    print("🎉 Excel链接更新完成！")
    print("=" * 70)
    
    success_count = sum(1 for _, success in results if success)
    if success_count == len(excel_files):
        print("✅ 所有文件更新成功！")
    else:
        print(f"⚠️  {success_count}/{len(excel_files)} 个文件更新成功")
    
    print("\n📁 输出文件：")
    print("   📊 新Excel文件：原文件名_GitHub版.xlsx")
    print("   📄 更新报告：Excel更新报告.md")
    print("   ⚙️  JSON报告：Excel更新报告.json")
    
    print("\n🚀 下一步操作：")
    print("1. 将新Excel文件分发给相关老师")
    print("2. 按照新链接格式写入手环")
    print("3. 更新平板浏览器的收藏链接")
    print("4. 进行测试验证")
    
    print("\n💡 提示：建议先更新5个手环进行小范围测试")

if __name__ == "__main__":
    main()